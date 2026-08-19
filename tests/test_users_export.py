"""Tests for CPQ user Excel export."""

from __future__ import annotations

import io

import httpx
import pytest
import respx
from openpyxl import load_workbook

from oracle_cpq_mcp.core.config import CPQProfile, CredentialSet
from oracle_cpq_mcp.core.cpq_client import CPQClient
from oracle_cpq_mcp.exporters.users_excel import (
    build_users_workbook,
    export_filename,
    fetch_all_users,
)


@pytest.fixture()
def profile() -> CPQProfile:
    return CPQProfile(
        customer_name="Test",
        customer_id="mycompany",
        environment="dev",
        base_url="https://dev.example.com",
        credentials=[CredentialSet(username="user", password="pass")],
        rest_version="v18",
    )


@pytest.fixture()
def client(profile: CPQProfile) -> CPQClient:
    return CPQClient(profile)


SAMPLE_USERS = [
    {
        "partyNumber": "BM_1",
        "login": "alice",
        "firstName": "Alice",
        "lastName": "Smith",
        "email": "alice@example.com",
        "status": {"value": 1, "displayValue": "Active"},
        "type": {"value": "FULL_ACCESS", "displayValue": "FullAccess"},
    },
    {
        "partyNumber": "BM_2",
        "login": "bob",
        "firstName": "Bob",
        "lastName": "Jones",
        "email": "bob@example.com",
        "status": {"value": 0, "displayValue": "Inactive"},
        "type": {"value": "FULL_ACCESS", "displayValue": "FullAccess"},
    },
]


def test_build_users_workbook_bytes_and_headers() -> None:
    xlsx_bytes = build_users_workbook(SAMPLE_USERS)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active

    assert sheet.title == "Users"
    assert [cell.value for cell in sheet[1]] == [
        "partyNumber",
        "login",
        "firstName",
        "lastName",
        "email",
        "status",
        "type",
        "language",
        "currency",
        "timeZone",
    ]
    assert sheet.max_row == 3
    assert sheet.cell(row=2, column=2).value == "alice"
    assert sheet.cell(row=2, column=6).value == "Active"
    assert sheet.cell(row=3, column=4).value == "Jones"


def test_build_users_workbook_custom_columns() -> None:
    xlsx_bytes = build_users_workbook(SAMPLE_USERS, columns=["login", "email"])
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active

    assert [cell.value for cell in sheet[1]] == ["login", "email"]
    assert sheet.cell(row=2, column=1).value == "alice"
    assert sheet.cell(row=2, column=2).value == "alice@example.com"


def test_export_filename() -> None:
    assert export_filename("mycompany", "dev") == "cpq_users_mycompany_dev.xlsx"


@respx.mock
def test_fetch_all_users_paginates(client: CPQClient) -> None:
    route = respx.get("https://dev.example.com/rest/v18/users").mock(
        side_effect=[
            httpx.Response(
                200,
                json={
                    "items": [SAMPLE_USERS[0]],
                    "hasMore": True,
                    "offset": 0,
                    "limit": 100,
                },
            ),
            httpx.Response(
                200,
                json={
                    "items": [SAMPLE_USERS[1]],
                    "hasMore": False,
                    "offset": 100,
                    "limit": 100,
                },
            ),
        ]
    )

    users = fetch_all_users(client)
    assert len(users) == 2
    assert users[0]["login"] == "alice"
    assert users[1]["login"] == "bob"
    assert route.call_count == 2
    assert "q" in route.calls[0].request.url.params
    assert "filter" not in route.calls[0].request.url.params
    assert route.calls[0].request.url.params["q"] == "{'status.value':{'$eq':1}}"


@respx.mock
def test_fetch_all_users_all_status_skips_q(client: CPQClient) -> None:
    route = respx.get("https://dev.example.com/rest/v18/users").mock(
        return_value=httpx.Response(
            200,
            json={"items": SAMPLE_USERS, "hasMore": False, "offset": 0, "limit": 100},
        )
    )

    users = fetch_all_users(client, status_filter="all")
    assert len(users) == 2
    assert "q" not in route.calls[0].request.url.params
